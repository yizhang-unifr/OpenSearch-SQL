#!/usr/bin/env python3
"""Rescore an existing pipeline run directory using the pre-computed gold SQL cache.

For every question result file (*_<db_id>.json) the script:
  1. Reads the PREDICTED_SQL from each evaluation stage (candidate_generate /
     align_correct / vote).
  2. Executes ONLY the predicted SQL against the database.
  3. Compares the result against the cached gold result set.
  4. Rewrites exec_res / exec_err / ves / GOLD_RESULT in the evaluation node
     (new key ``gold_from_cache`` is also added; all existing keys are kept).
  5. Writes updated files to a NEW sibling directory named
     <original_dir>_rescored_<YYYYMMDD>.

A fresh -statistics.json and XLSX report are generated in the new directory.
A side-by-side comparison sheet is included in the XLSX.

Validation requirements for an input run directory
---------------------------------------------------
  * Must contain -statistics.json
  * Must contain at least one *_<db_id>.json file
  * logs/ sub-directory is ignored

Usage
-----
  uv run run/rescore_with_gold_cache.py \\
      --run-dir results/test_data_point/no_few_shot/baseline/pipe9_xxx/2026-05-19-00-04-00

  # Specify cache explicitly
  uv run run/rescore_with_gold_cache.py \\
      --run-dir <path> --cache-path data/gold_sql_cache/test_data_point_meteo_gold_sql.json

  # Dry run (print what would change, no files written)
  uv run run/rescore_with_gold_cache.py --run-dir <path> --dry-run
"""

from __future__ import annotations

import argparse
import copy
import json
import logging
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any

# ---------------------------------------------------------------------------
# Path bootstrap
# ---------------------------------------------------------------------------
_SCRIPT_DIR   = Path(__file__).resolve().parent          # run/
_OPENSEARCH   = _SCRIPT_DIR.parent                       # src/OpenSearch-SQL/
_PROJECT_ROOT = _OPENSEARCH.parents[1]                   # ontology_retriever/
_SRC          = _OPENSEARCH / "src"
sys.path.insert(0, str(_SRC))

try:
    from dotenv import load_dotenv
    load_dotenv(_PROJECT_ROOT / ".env", override=False)
except ImportError:
    pass

from runner.execution import compare_with_cached_gold, _normalize_sql
from runner.gold_sql_cache import GoldSqlCache
from runner.statistics_manager import Statistics, StatisticsManager

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

_PRED_TIMEOUT_S = 600   # 10 min for predicted SQL re-execution

EVAL_STAGES = ["candidate_generate", "align_correct", "vote"]

# ---------------------------------------------------------------------------
# Helpers: run-dir validation
# ---------------------------------------------------------------------------

def validate_run_dir(run_dir: Path, db_id: str) -> list[Path]:
    """Return sorted list of *_<db_id>.json files; raise if dir is invalid."""
    if not run_dir.is_dir():
        raise FileNotFoundError(f"Not a directory: {run_dir}")
    if not (run_dir / "-statistics.json").exists():
        raise ValueError(f"No -statistics.json found in {run_dir}")
    files = sorted(
        run_dir.glob(f"*_{db_id}.json"),
        key=lambda p: int(p.stem.split("_")[0]) if p.stem.split("_")[0].isdigit() else -1,
    )
    if not files:
        raise ValueError(f"No *_{db_id}.json files found in {run_dir}")
    return files


# ---------------------------------------------------------------------------
# Single-question rescoring
# ---------------------------------------------------------------------------

def _rescore_question(
    nodes: list[dict],
    cache: GoldSqlCache,
    q_id: int,
    dry_run: bool,
) -> tuple[list[dict], dict]:
    """Return (updated_nodes, delta) where delta maps stage → (old_exec, new_exec)."""
    new_nodes  = copy.deepcopy(nodes)
    delta: dict[str, dict] = {}

    # Locate evaluation node
    eval_idx = next(
        (i for i, n in enumerate(new_nodes) if n.get("node_type") == "evaluation"),
        None,
    )
    if eval_idx is None:
        return new_nodes, delta

    ev_node = new_nodes[eval_idx]

    if not cache.has(q_id):
        log.debug("  q_id=%d: not in cache, skipping", q_id)
        return new_nodes, delta

    cached_entry = cache.get_entry(q_id)
    if cached_entry is None or cached_entry.get("status") != "success":
        log.debug("  q_id=%d: cache status=%s, skipping",
                  q_id, cached_entry and cached_entry.get("status"))
        return new_nodes, delta

    gold_set = cache.get_gold_set(q_id)
    t_gold   = cache.get_duration(q_id) or 0.0
    gold_result_list = cached_entry.get("result")   # list of lists

    if gold_set is None:
        return new_nodes, delta

    for stage in EVAL_STAGES:
        stage_data = ev_node.get(stage)
        if stage_data is None:
            continue

        predicted_sql = stage_data.get("PREDICTED_SQL", "--")
        old_exec_res  = stage_data.get("exec_res")
        old_ves       = stage_data.get("ves", 0.0)

        if dry_run:
            log.info("  [DRY] q_id=%d  stage=%-20s  old_exec=%s",
                     q_id, stage, old_exec_res)
            delta[stage] = {"old_exec_res": old_exec_res, "new_exec_res": "(dry)"}
            continue

        t0 = time.perf_counter()
        response = compare_with_cached_gold(
            predicted_sql=predicted_sql,
            gold_set=gold_set,
            t_gold=t_gold,
            meta_time_out=_PRED_TIMEOUT_S,
        )
        elapsed = time.perf_counter() - t0

        new_exec_res = response["exec_res"]
        new_exec_err = response["exec_err"]
        new_ves      = response["ves"]

        # Update existing keys (keep all other keys intact)
        stage_data["exec_res"]       = new_exec_res
        stage_data["exec_err"]       = new_exec_err
        stage_data["ves"]            = new_ves
        stage_data["GOLD_RESULT"]    = gold_result_list
        # New key only:
        stage_data["gold_from_cache"] = True

        delta[stage] = {
            "old_exec_res": old_exec_res,
            "old_ves":      old_ves,
            "new_exec_res": new_exec_res,
            "new_ves":      new_ves,
            "elapsed_s":    round(elapsed, 2),
        }
        log.debug(
            "  q_id=%d  stage=%-20s  exec %s→%s  ves %.3f→%.3f  %.1fs",
            q_id, stage, old_exec_res, new_exec_res, old_ves, new_ves, elapsed,
        )

    return new_nodes, delta


# ---------------------------------------------------------------------------
# Statistics rebuild
# ---------------------------------------------------------------------------

def _rebuild_statistics(run_dir: Path, db_id: str) -> Statistics:
    """Recompute Statistics from rescored JSON files in *run_dir*."""
    stats = Statistics()
    for jf in sorted(run_dir.glob(f"*_{db_id}.json"),
                     key=lambda p: int(p.stem.split("_")[0]) if p.stem.split("_")[0].isdigit() else -1):
        if jf.name.startswith("-"):
            continue
        parts = jf.stem.split("_", 1)
        if len(parts) != 2 or not parts[0].isdigit():
            continue
        q_id_str, d_id = parts
        with jf.open(encoding="utf-8") as f:
            nodes = json.load(f)
        ev_node = next((n for n in nodes if n.get("node_type") == "evaluation"), None)
        if ev_node is None:
            continue
        for stage in EVAL_STAGES:
            sd = ev_node.get(stage)
            if sd is None:
                continue
            stats.total[stage]   = stats.total.get(stage, 0) + 1
            stats.ves_sum[stage] = stats.ves_sum.get(stage, 0.0) + float(sd.get("ves", 0.0))
            exec_res = sd.get("exec_res")
            exec_err = sd.get("exec_err", "")
            if exec_res == 1:
                stats.corrects.setdefault(stage, []).append((d_id, q_id_str))
            elif exec_err == "incorrect answer":
                stats.incorrects.setdefault(stage, []).append((d_id, q_id_str))
            else:
                stats.errors.setdefault(stage, []).append((d_id, q_id_str, str(exec_err)))
    return stats


# ---------------------------------------------------------------------------
# XLSX comparison report
# ---------------------------------------------------------------------------

def _export_comparison_xlsx(
    original_dir: Path,
    rescored_dir: Path,
    db_id: str,
    out_path: Path,
) -> None:
    """Build an XLSX with before/after comparison."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — skipping XLSX export")
        return

    _GREEN  = "C6EFCE"
    _RED    = "FFC7CE"
    _YELLOW = "FFEB9C"
    _BLUE   = "BDD7EE"
    _GREY   = "F2F2F2"
    _BOLD   = Font(bold=True)
    _HFILL  = PatternFill("solid", fgColor=_BLUE)
    _TOP    = Alignment(vertical="top")

    def _fill(exec_res):
        if exec_res == 1:   return PatternFill("solid", fgColor=_GREEN)
        if exec_res == 0:   return PatternFill("solid", fgColor=_RED)
        return PatternFill("solid", fgColor=_YELLOW)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Comparison sheet ────────────────────────────────────────────────────
    ws = wb.create_sheet("Comparison")
    headers = [
        "Q_ID", "Question", "Stage",
        "OLD exec_res", "OLD VES",
        "NEW exec_res", "NEW VES",
        "Changed", "Gold SQL", "Predicted SQL",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = _BOLD
        c.fill = _HFILL
        c.alignment = _TOP

    row = 2
    orig_files  = {
        int(p.stem.split("_")[0]): p
        for p in original_dir.glob(f"*_{db_id}.json")
        if not p.name.startswith("-") and p.stem.split("_")[0].isdigit()
    }
    new_files   = {
        int(p.stem.split("_")[0]): p
        for p in rescored_dir.glob(f"*_{db_id}.json")
        if not p.name.startswith("-") and p.stem.split("_")[0].isdigit()
    }

    for q_id in sorted(orig_files):
        with orig_files[q_id].open() as f:
            orig_nodes = json.load(f)
        new_nodes = json.load(new_files[q_id].open()) if q_id in new_files else orig_nodes

        orig_ev = next((n for n in orig_nodes if n.get("node_type") == "evaluation"), {})
        new_ev  = next((n for n in new_nodes  if n.get("node_type") == "evaluation"), {})

        question = (orig_ev.get("vote") or orig_ev.get("candidate_generate") or {}).get("Question", "")
        gold_sql = (orig_ev.get("vote") or orig_ev.get("candidate_generate") or {}).get("GOLD_SQL", "")

        bg = PatternFill("solid", fgColor=_GREY) if row % 2 == 0 else None

        for stage in EVAL_STAGES:
            od = orig_ev.get(stage, {})
            nd = new_ev.get(stage, {})
            old_exec = od.get("exec_res")
            new_exec = nd.get("exec_res")
            changed  = old_exec != new_exec

            vals = [
                q_id, question, stage,
                old_exec, round(od.get("ves", 0.0), 4),
                new_exec, round(nd.get("ves", 0.0), 4),
                changed,
                gold_sql[:120],
                nd.get("PREDICTED_SQL", "")[:120],
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.alignment = _TOP
                if col == 4: c.fill = _fill(old_exec)
                if col == 6: c.fill = _fill(new_exec)
                elif bg and col not in (4, 6): c.fill = bg
            row += 1

    col_widths = {1: 6, 2: 40, 3: 20, 4: 12, 5: 8, 6: 12, 7: 8, 8: 8, 9: 60, 10: 60}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "D2"

    # ── Summary comparison ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Score Delta")
    old_stats_path = original_dir / "-statistics.json"
    new_stats_path = rescored_dir  / "-statistics.json"

    ws2.append(["Stage", "Metric", "Original", "Rescored", "Delta"])
    for c in ws2[1]:
        c.font = _BOLD
        c.fill = _HFILL

    def _read_stats(path: Path) -> dict:
        if path.exists():
            with path.open() as f:
                return json.load(f).get("counts", {})
        return {}

    old_counts = _read_stats(old_stats_path)
    new_counts = _read_stats(new_stats_path)

    for stage in EVAL_STAGES:
        oc = old_counts.get(stage, {})
        nc = new_counts.get(stage, {})
        for metric in ("EX", "VES", "correct", "error"):
            ov = oc.get(metric, 0)
            nv = nc.get(metric, 0)
            delta = round(nv - ov, 4) if isinstance(ov, (int, float)) else ""
            row_vals = [stage, metric, ov, nv, delta]
            ws2.append(row_vals)
            if metric in ("EX", "VES") and isinstance(delta, float):
                c = ws2.cell(row=ws2.max_row, column=5)
                c.fill = PatternFill("solid", fgColor=_GREEN if delta > 0 else (_RED if delta < 0 else "FFFFFF"))

    for col, w in {1: 22, 2: 8, 3: 10, 4: 10, 5: 10}.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Also export rescored full results using existing exporter ──────────
    try:
        sys.path.insert(0, str(_OPENSEARCH / "eval"))
        from export_results_xlsx import export as _export_full
        full_xlsx = out_path.parent / out_path.name.replace("_comparison", "_full")
        _export_full(rescored_dir, full_xlsx)
        log.info("Full results XLSX → %s", full_xlsx)
    except Exception as exc:
        log.warning("Could not export full XLSX: %s", exc)

    wb.save(out_path)
    log.info("Comparison XLSX → %s", out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args(argv=None):
    p = argparse.ArgumentParser(
        description="Rescore a pipeline run using the pre-computed gold SQL cache.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--run-dir", required=True,
                   help="Path to the existing pipeline run directory.")
    p.add_argument("--db-id", default="meteo",
                   help="Database identifier (default: meteo)")
    p.add_argument("--dataset", default="test_data_point",
                   help="Dataset name for locating the default cache (default: test_data_point)")
    p.add_argument("--cache-path", default=None,
                   help="Path to gold SQL cache JSON (default: auto-detected)")
    p.add_argument("--out-dir", default=None,
                   help="Output directory (default: <run_dir>_rescored_<YYYYMMDD>)")
    p.add_argument("--dry-run", action="store_true",
                   help="Show what would change without writing any files")
    p.add_argument("--no-xlsx", action="store_true",
                   help="Skip XLSX report generation")
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    run_dir = Path(args.run_dir).resolve()

    # Validate input
    try:
        question_files = validate_run_dir(run_dir, args.db_id)
    except (FileNotFoundError, ValueError) as e:
        log.error("%s", e)
        return 1

    log.info("Input run dir : %s", run_dir)
    log.info("Questions     : %d", len(question_files))

    # Locate cache
    if args.cache_path:
        cache_path = Path(args.cache_path)
    else:
        cache_path = GoldSqlCache.cache_path(_PROJECT_ROOT, args.dataset, args.db_id)

    if not cache_path.exists():
        log.error("Gold SQL cache not found: %s — run run_gold_sql.py first.", cache_path)
        return 1

    cache = GoldSqlCache(cache_path)
    log.info("Cache loaded  : %s (%d entries, %d success)",
             cache_path,
             len(cache._raw),
             sum(1 for e in cache._raw.values() if e.get("status") == "success"))

    # Output directory
    if args.out_dir:
        out_dir = Path(args.out_dir)
    else:
        ts      = datetime.now().strftime("%Y%m%d")
        out_dir = run_dir.parent / f"{run_dir.name}_rescored_{ts}"

    if not args.dry_run:
        out_dir.mkdir(parents=True, exist_ok=True)
        log.info("Output dir    : %s", out_dir)

    # Copy non-question files first (args, statistics placeholder, etc.)
    if not args.dry_run:
        for f in run_dir.iterdir():
            if f.name.startswith("-") and f.suffix == ".json" and f.name != "-statistics.json":
                import shutil
                shutil.copy2(f, out_dir / f.name)

    # Rescore each question
    total_changed = 0
    t_wall = time.perf_counter()

    for i, qf in enumerate(question_files, 1):
        parts = qf.stem.split("_", 1)
        q_id  = int(parts[0])

        with qf.open(encoding="utf-8") as f:
            nodes = json.load(f)

        new_nodes, delta = _rescore_question(nodes, cache, q_id, args.dry_run)

        changed = any(
            d.get("old_exec_res") != d.get("new_exec_res")
            for d in delta.values()
            if "old_exec_res" in d and "new_exec_res" in d
        )
        if changed:
            total_changed += 1

        if not args.dry_run:
            out_path = out_dir / qf.name
            with out_path.open("w", encoding="utf-8") as f:
                json.dump(new_nodes, f, ensure_ascii=False, default=str)

        if i % 50 == 0 or i == len(question_files):
            elapsed = time.perf_counter() - t_wall
            log.info("  [%d/%d]  changed_so_far=%d  elapsed=%.0fs",
                     i, len(question_files), total_changed, elapsed)

    log.info("Rescored %d/%d questions changed.", total_changed, len(question_files))

    if args.dry_run:
        return 0

    # Rebuild statistics
    stats = _rebuild_statistics(out_dir, args.db_id)
    stats_path = out_dir / "-statistics.json"
    with stats_path.open("w", encoding="utf-8") as f:
        json.dump(stats.to_dict(), f, indent=4, ensure_ascii=False)
    log.info("Statistics written → %s", stats_path)

    old_counts = json.loads((run_dir / "-statistics.json").read_text())["counts"]
    new_counts = stats.to_dict()["counts"]
    log.info("Score delta (vote stage):")
    for metric in ("EX", "VES"):
        ov = old_counts.get("vote", {}).get(metric, 0)
        nv = new_counts.get("vote", {}).get(metric, 0)
        log.info("  vote.%-6s  %.4f → %.4f  (Δ%+.4f)", metric, ov, nv, nv - ov)

    if not args.no_xlsx:
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx = out_dir / f"results_comparison_{ts}.xlsx"
        _export_comparison_xlsx(run_dir, out_dir, args.db_id, xlsx)

    return 0


if __name__ == "__main__":
    sys.exit(main())
