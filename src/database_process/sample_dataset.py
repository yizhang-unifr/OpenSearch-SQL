"""Sample questions per category from an XLSX file into pipeline-ready JSON.

Input file resolution:
  filename only   →  <project_root>/data/<filename>
  relative path   →  relative to cwd
  absolute path   →  used as-is

Output files are written to:
  src/OpenSearch-SQL/data/data_preprocess/<output_file>[_point|_bbox].json

Usage:
  uv run src/database_process/sample_dataset.py \\
      --input-file Thessaly_NOA.xlsx \\
      --output-file thesaly_sample \\
      --mode both

  uv run src/database_process/sample_dataset.py \\
      --input-file Thessaly_NOA.xlsx \\
      --output-file heat_wave_point \\
      --mode point --n 3
"""

from __future__ import annotations

import argparse
import json
import random
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

_SCRIPT_DIR      = Path(__file__).resolve().parent
_OPENSEARCH_ROOT = _SCRIPT_DIR.parents[1]          # src/OpenSearch-SQL/
_PROJECT_DATA    = _OPENSEARCH_ROOT.parents[1] / "data"   # ontology_retriever/data/
_OUT_DIR         = _OPENSEARCH_ROOT / "data" / "data_preprocess"


def _resolve_input(input_file: str) -> Path:
    p = Path(input_file)
    if p.is_absolute():
        return p
    if p.exists():
        return p.resolve()
    candidate = _PROJECT_DATA / p
    if candidate.exists():
        return candidate
    raise FileNotFoundError(
        f"Cannot find '{input_file}'. "
        f"Tried: {p.resolve()} and {candidate}"
    )


def load_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        record["_row"] = row_idx
        rows.append(record)
    return rows


def sample_by_category(rows: list[dict], n: int, seed: int,
                        category_col: str) -> list[dict]:
    rng = random.Random(seed)
    by_cat: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        cat = str(r.get(category_col) or "").strip()
        if cat:
            by_cat[cat].append(r)

    sampled: list[dict] = []
    for cat in sorted(by_cat):
        chosen = rng.sample(by_cat[cat], min(n, len(by_cat[cat])))
        sampled.extend(chosen)
    return sampled


def build_records(sampled: list[dict], sql_col: str, question_col: str,
                  category_col: str, geo_mode: str) -> list[dict]:
    records = []
    for i, r in enumerate(sampled):
        sql = r.get(sql_col) or ""
        records.append({
            "question_id": i,
            "question":    str(r.get(question_col) or "").strip(),
            "db_id":       "meteo",
            "evidence":    "",
            "SQL":         sql.strip() if isinstance(sql, str) else "",
            "category":    str(r.get(category_col) or "").strip(),
            "source_line_no": r["_row"],
            "geo_filter_mode": geo_mode,
        })
    return records


def write_json(records: list[dict], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(records, indent=4, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"  Wrote {len(records)} records → {out_path}")


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Sample questions per category from an XLSX into pipeline JSON.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("--input-file", required=True,
                    help="XLSX filename (looked up in data/) or path")
    ap.add_argument("--output-file", required=True,
                    help="Output stem; produces <stem>_point.json and/or <stem>_bbox.json")
    ap.add_argument("--mode", default="both", choices=["point", "bbox", "both"],
                    help="Geo filter mode to generate (default: both)")
    ap.add_argument("--n", type=int, default=2,
                    help="Samples per category (default: 2)")
    ap.add_argument("--seed", type=int, default=42,
                    help="Random seed (default: 42)")
    # Column name overrides
    ap.add_argument("--question-col", default="question_generated",
                    help="Question column name (default: question_generated)")
    ap.add_argument("--category-col", default="category",
                    help="Category column name (default: category)")
    ap.add_argument("--point-sql-col", default="point_sql",
                    help="Point SQL column name (default: point_sql)")
    ap.add_argument("--bbox-sql-col", default="bbox_sql",
                    help="Bbox SQL column name (default: bbox_sql)")
    args = ap.parse_args()

    xlsx_path = _resolve_input(args.input_file)
    print(f"Loading {xlsx_path} …")
    rows = load_xlsx(xlsx_path)

    cat_counts: Counter = Counter(
        str(r.get(args.category_col) or "").strip() for r in rows
    )
    print(f"  {len(rows)} rows, {len(cat_counts)} categories: "
          + ", ".join(f"{c}({n})" for c, n in sorted(cat_counts.items()) if c))

    sampled = sample_by_category(rows, args.n, args.seed, args.category_col)
    print(f"\nSampled {len(sampled)} questions ({args.n}/category, seed={args.seed})")
    breakdown = Counter(str(r.get(args.category_col, "")).strip() for r in sampled)
    for cat in sorted(breakdown):
        print(f"  {cat}: {breakdown[cat]}")

    stem = args.output_file
    if args.mode in ("point", "both"):
        print("\nWriting point variant …")
        records = build_records(sampled, args.point_sql_col,
                                args.question_col, args.category_col, "points")
        write_json(records, _OUT_DIR / f"{stem}_point.json")

    if args.mode in ("bbox", "both"):
        print("\nWriting bbox variant …")
        records = build_records(sampled, args.bbox_sql_col,
                                args.question_col, args.category_col, "bbox")
        write_json(records, _OUT_DIR / f"{stem}_bbox.json")

    print("\nDone.")


if __name__ == "__main__":
    main()
