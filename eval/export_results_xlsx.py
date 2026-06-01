"""Export pipeline run results to XLSX.

Usage:
    python eval/export_results_xlsx.py --run_dir results/landcover10/no_few_shot/full/pipe9_xxx/2026-05-11-08-30-00
    python eval/export_results_xlsx.py --results_root results/landcover10  # picks latest run
    python eval/export_results_xlsx.py --run_dir <path> --output my_results.xlsx
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_GREEN  = "C6EFCE"   # correct
_RED    = "FFC7CE"   # incorrect / error
_YELLOW = "FFEB9C"   # warning / partial
_BLUE   = "BDD7EE"   # header
_GREY   = "F2F2F2"   # alternate row


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _find_latest_run(results_root: Path) -> Path:
    """Walk results_root recursively and return the most-recently-modified
    directory that contains at least one *_meteo.json file."""
    candidates = []
    for p in results_root.rglob("*_meteo.json"):
        candidates.append(p.parent)
    if not candidates:
        raise FileNotFoundError(f"No *_meteo.json files under {results_root}")
    return max(set(candidates), key=lambda d: d.stat().st_mtime)


def _load_run(run_dir: Path) -> List[Dict[str, Any]]:
    """Return a list of question records sorted by question_id."""
    records = []
    for json_file in sorted(run_dir.glob("*_*.json")):
        if json_file.name.startswith("-"):
            continue
        parts = json_file.stem.split("_", 1)
        if len(parts) != 2:
            continue
        q_id_str, db_id = parts
        try:
            q_id = int(q_id_str)
        except ValueError:
            continue

        with json_file.open(encoding="utf-8") as f:
            nodes: List[Dict] = json.load(f)

        record = {"q_id": q_id, "db_id": db_id, "nodes": {}, "timings": {}, "gen_stats": None}
        for node in nodes:
            nt = node.get("node_type", "")
            record["nodes"][nt] = node
            if "duration_s" in node:
                record["timings"][nt] = node["duration_s"]
            if nt == "candidate_generate" and node.get("generation_stats"):
                record["gen_stats"] = node["generation_stats"]

        records.append(record)

    return sorted(records, key=lambda r: r["q_id"])


def _get_eval(record: Dict, eval_node: str = "vote") -> Dict:
    ev = record["nodes"].get("evaluation", {})
    return ev.get(eval_node, {})


def _fmt_result(val) -> str:
    if val is None:
        return ""
    if isinstance(val, list):
        return json.dumps(val, ensure_ascii=False)
    return str(val)


def _exec_label(exec_res) -> str:
    if exec_res == 1:
        return "correct"
    if exec_res == 0:
        return "incorrect"
    return str(exec_res) if exec_res is not None else "error"


def _exec_fill(exec_res) -> Optional[PatternFill]:
    if exec_res == 1:
        return PatternFill("solid", fgColor=_GREEN)
    if exec_res == 0:
        return PatternFill("solid", fgColor=_RED)
    return PatternFill("solid", fgColor=_YELLOW)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor=_BLUE)
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP  = Alignment(vertical="top")

EVAL_NODES = ["candidate_generate", "align_correct", "vote"]
PIPELINE_NODES = [
    "generate_db_schema",
    "extract_col_value",
    "extract_query_noun",
    "column_retrieve_and_other_info",
    "implicit_context_enhance",
    "candidate_generate",
    "align_correct",
    "vote",
    "evaluation",
]


def _write_header(ws, headers: List[str], row: int = 1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _TOP


def _set_col_widths(ws, widths: Dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_summary_sheet(wb: openpyxl.Workbook, records: List[Dict]):
    """Sheet 1 – one row per question, evaluation results + node timings."""
    ws = wb.create_sheet("Summary")

    headers = ["Q_ID", "DB", "Question"]
    for en in EVAL_NODES:
        label = en.replace("_", " ").title()
        headers += [f"{label} | Correct", f"{label} | VES", f"{label} | Error"]
    for pn in PIPELINE_NODES:
        headers.append(f"⏱ {pn} (s)")
    # candidate_generate per-phase LLM stats
    headers += [
        "CG calls",
        "CG total (ms)", "CG min (ms)", "CG max (ms)", "CG avg (ms)",
        "CG in tok", "CG out tok",
    ]

    _write_header(ws, headers)

    for r_idx, rec in enumerate(records, 2):
        ev_vote = _get_eval(rec, "vote")
        col = 1
        fill = PatternFill("solid", fgColor=_GREY) if r_idx % 2 == 0 else None

        def _w(val, fill_override=None):
            nonlocal col
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.alignment = _TOP
            if fill_override:
                cell.fill = fill_override
            elif fill:
                cell.fill = fill
            col += 1

        _w(rec["q_id"])
        _w(rec["db_id"])
        _w(ev_vote.get("Question", ""))

        for en in EVAL_NODES:
            ev = _get_eval(rec, en)
            exec_res = ev.get("exec_res")
            _w(_exec_label(exec_res), fill_override=_exec_fill(exec_res))
            _w(round(ev.get("ves", 0.0), 4) if ev else "")
            err = ev.get("exec_err", "")
            _w("" if err in ("incorrect answer", None) else str(err)[:120])

        for pn in PIPELINE_NODES:
            _w(rec["timings"].get(pn, ""))

        gs = rec.get("gen_stats") or {}
        _w(gs.get("n_calls", ""))
        _w(gs.get("total_duration_ms", ""))
        _w(gs.get("min_duration_ms", ""))
        _w(gs.get("max_duration_ms", ""))
        _w(gs.get("avg_duration_ms", ""))
        _w(gs.get("total_input_tokens", ""))
        _w(gs.get("total_output_tokens", ""))

    _set_col_widths(ws, {1: 6, 2: 8, 3: 45, **{c: 12 for c in range(4, len(headers) + 1)}})
    ws.freeze_panes = "D2"


def build_sql_sheet(wb: openpyxl.Workbook, records: List[Dict]):
    """Sheet 2 – full SQL text and execution results for comparison."""
    ws = wb.create_sheet("SQL & Results")

    headers = [
        "Q_ID", "Question",
        "Node", "Exec", "VES",
        "Gold SQL", "Predicted SQL",
        "Gold Result", "Predicted Result",
        "Exec Error",
    ]
    _write_header(ws, headers)

    row = 2
    for rec in records:
        ev_vote = _get_eval(rec, "vote")
        question = ev_vote.get("Question", "")
        gold_sql  = ev_vote.get("GOLD_SQL", "")
        gold_res  = _fmt_result(ev_vote.get("GOLD_RESULT"))

        for en in EVAL_NODES:
            ev = _get_eval(rec, en)
            if not ev:
                continue
            exec_res = ev.get("exec_res")
            fill = _exec_fill(exec_res)
            values = [
                rec["q_id"],
                question,
                en,
                _exec_label(exec_res),
                round(ev.get("ves", 0.0), 4),
                gold_sql,
                ev.get("PREDICTED_SQL", ""),
                gold_res,
                _fmt_result(ev.get("PREDICTED_RESULT")),
                ev.get("exec_err", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = _WRAP
                if col == 4:   # Exec column gets colour
                    cell.fill = fill
            row += 1

    _set_col_widths(ws, {1: 6, 2: 35, 3: 18, 4: 12, 5: 8,
                         6: 60, 7: 60, 8: 30, 9: 30, 10: 40})
    ws.freeze_panes = "C2"
    for r in range(2, row):
        ws.row_dimensions[r].height = 90


def build_timing_sheet(wb: openpyxl.Workbook, records: List[Dict]):
    """Sheet 3 – plugin and node timing breakdown."""
    ws = wb.create_sheet("Timing")

    headers = ["Q_ID", "Question"] + [f"{pn} (s)" for pn in PIPELINE_NODES]
    _write_header(ws, headers)

    for r_idx, rec in enumerate(records, 2):
        ev_vote = _get_eval(rec, "vote")
        question = ev_vote.get("Question", "")
        fill = PatternFill("solid", fgColor=_GREY) if r_idx % 2 == 0 else None

        ws.cell(row=r_idx, column=1, value=rec["q_id"]).alignment = _TOP
        ws.cell(row=r_idx, column=2, value=question).alignment = _TOP
        if fill:
            ws.cell(row=r_idx, column=1).fill = fill
            ws.cell(row=r_idx, column=2).fill = fill

        for col, pn in enumerate(PIPELINE_NODES, 3):
            cell = ws.cell(row=r_idx, column=col, value=rec["timings"].get(pn, ""))
            cell.alignment = _TOP
            if fill:
                cell.fill = fill

    _set_col_widths(ws, {1: 6, 2: 40, **{c: 14 for c in range(3, len(headers) + 1)}})
    ws.freeze_panes = "C2"


def build_review_sheet(wb: openpyxl.Workbook, records: List[Dict]):
    """Sheet 4 – LLM constraint review trace per candidate.

    Handles both the new LLM-review format and the legacy plugin-registry format
    so old result directories can still be exported.
    """
    ws = wb.create_sheet("Review Trace")

    headers = ["Q_ID", "Candidate #", "Reviewer", "Changed", "Issues / Warnings", "Details"]
    _write_header(ws, headers)

    row = 2
    for rec in records:
        cg = rec["nodes"].get("candidate_generate", {})
        trace = cg.get("plugin_trace", [])
        if not trace:
            continue

        first = trace[0] if trace else {}

        if "review" in first:
            # New format: flat list of per-candidate review dicts
            for entry in trace:
                issues = entry.get("issues") or []
                err = entry.get("error", "")
                detail = err if err else json.dumps(entry.get("raw", ""), ensure_ascii=False)[:200]
                changed = entry.get("changed", False)
                fill = PatternFill("solid", fgColor=_YELLOW) if changed else None
                values = [
                    rec["q_id"],
                    entry.get("candidate_index", "?"),
                    entry.get("review", "llm_constraint_review"),
                    changed,
                    "; ".join(issues) if issues else "",
                    detail,
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.alignment = _TOP
                    if fill and col >= 3:
                        cell.fill = fill
                row += 1
        else:
            # Legacy plugin-registry format
            for cand in trace:
                if not isinstance(cand, dict):
                    continue
                cand_idx = cand.get("candidate_index", "?")
                for entry in cand.get("trace", []):
                    changed = entry.get("changed", False)
                    fill = PatternFill("solid", fgColor=_YELLOW) if changed else None
                    values = [
                        rec["q_id"],
                        cand_idx,
                        entry.get("plugin", ""),
                        changed,
                        "; ".join(entry.get("warnings", [])),
                        json.dumps(entry.get("metadata", {}), ensure_ascii=False)[:200],
                    ]
                    for col, val in enumerate(values, 1):
                        cell = ws.cell(row=row, column=col, value=val)
                        cell.alignment = _TOP
                        if fill and col >= 3:
                            cell.fill = fill
                    row += 1

    _set_col_widths(ws, {1: 6, 2: 12, 3: 28, 4: 10, 5: 45, 6: 50})
    ws.freeze_panes = "C2"


def build_llm_calls_sheet(wb: openpyxl.Workbook, records: List[Dict]):
    """Sheet – one row per individual candidate_generate LLM call.

    Breaks the per-phase generation_stats down to per-call duration + token counts.
    """
    ws = wb.create_sheet("LLM Calls")

    headers = ["Q_ID", "DB", "Call #", "Duration (ms)", "Input tokens", "Output tokens"]
    _write_header(ws, headers)

    row = 2
    for rec in records:
        gs = rec.get("gen_stats") or {}
        durs = gs.get("per_call_duration_ms") or []
        ins  = gs.get("input_tokens") or []
        outs = gs.get("output_tokens") or []
        n = max(len(durs), len(ins), len(outs))
        for i in range(n):
            values = [
                rec["q_id"],
                rec["db_id"],
                i + 1,
                durs[i] if i < len(durs) else "",
                ins[i] if i < len(ins) else "",
                outs[i] if i < len(outs) else "",
            ]
            for col, val in enumerate(values, 1):
                ws.cell(row=row, column=col, value=val).alignment = _TOP
            row += 1

    _set_col_widths(ws, {1: 6, 2: 8, 3: 8, 4: 14, 5: 14, 6: 14})
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def export(run_dir: Path, output_path: Optional[Path] = None) -> Path:
    records = _load_run(run_dir)
    if not records:
        raise ValueError(f"No question JSON files found in {run_dir}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    build_summary_sheet(wb, records)
    build_sql_sheet(wb, records)
    build_timing_sheet(wb, records)
    build_review_sheet(wb, records)
    build_llm_calls_sheet(wb, records)

    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = run_dir / f"results_{ts}.xlsx"

    wb.save(output_path)
    return output_path


def main():
    ap = argparse.ArgumentParser(description="Export pipeline run results to XLSX")
    ap.add_argument(
        "--run_dir",
        help="Path to a specific run directory (contains *_meteo.json files)",
    )
    ap.add_argument(
        "--results_root",
        default="results",
        help="Root results directory; latest run is picked automatically (default: results/)",
    )
    ap.add_argument("--output", help="Output .xlsx path (default: <run_dir>/results_<ts>.xlsx)")
    args = ap.parse_args()

    run_dir = Path(args.run_dir) if args.run_dir else _find_latest_run(Path(args.results_root))
    output_path = Path(args.output) if args.output else None

    out = export(run_dir, output_path)
    print(f"Exported → {out}")


if __name__ == "__main__":
    main()
